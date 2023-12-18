# Import Packages
from flask import Flask, render_template, send_file
from flask_wtf import FlaskForm
from wtforms import FileField, SubmitField
from werkzeug.utils import secure_filename
import os
from wtforms.validators import InputRequired
import openpyxl
from openpyxl.chart import BarChart, Reference
import urllib.request
import json
import pandas as pd
import shutil

# Assign global variables
filename = ''
file_uploaded = False
ALLOWED_EXTENSIONS = set(['XLSX','xlsx','xls','XLS'])

def allowed_file(filename):
	""" If the extension is allowed the function check it """
	return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Create app in flask
app = Flask(__name__)
app.config['SECRET_KEY'] = 'supersecretkey'
app.config['UPLOAD_FOLDER'] = 'static/upload_excel'
app.config['DOWNLOAD_FOLDER'] = 'static/download_excel'

class UploadFileForm(FlaskForm):
	""" It validates the file """
	file = FileField("File", validators = [InputRequired()])
	submit = SubmitField("Upload File")

@app.route('/', methods=["GET", "POST"])
@app.route('/home', methods=["GET", "POST"])
def home():
	""" Renders the home page """
	global filename, file_uploaded
	form = UploadFileForm()
	file_uploaded = False
	if form.validate_on_submit():
		file = form.file.data # Grab the file				
		file.save(os.path.join(os.path.abspath(os.path.dirname(__file__)), app.config['UPLOAD_FOLDER'], secure_filename(file.filename))) # Save the file
		filename = secure_filename(file.filename) # Get the file name
		original = f'static/upload_excel/{filename}' # Original file location
		target = f'static/download_excel/{filename}' # Modified file location
		shutil.copyfile(original, target) # Copy file to destination location
		excel_data(filename) # Call the function excel_data
		
		file_uploaded = True
		return render_template("index.html", form=form, file_uploaded = file_uploaded)
	return render_template("index.html", form=form, file_uploaded = file_uploaded)
	
@app.route('/download', methods=["GET", "POST"])	
def download_file():
	""" Downloads the modified file """
	global filename
	path = f"static/download_excel/{filename}"
	return send_file(path, as_attachment = True)


def weather(city):
	""" Gets the weather data from API """
	city = city.replace(" ", "%20")
	url = f'https://api.openweathermap.org/data/2.5/weather?q={city.lower()}&appid=778fedad8e7d9547b3eae10f40c89437'
	response = urllib.request.urlopen(url) #Opening url in python
	result = json.loads(response.read()) #Loading json in python
	city_dict = {}
	city_dict["Temperature (Celsius)"] = round(result['main']['temp'] - 273.15, 2)
	city_dict["Minimum Temperature (Celsius)"] = round(result['main']['temp_min'] - 273.15, 2)
	city_dict["Maximum Temperature (Celsius)"] = round(result['main']['temp_max'] - 273.15, 2)
	city_dict["Feels like (Celsius)"] = round(result['main']['feels_like'] - 273.15, 2)
	city_dict["City Name"] = city.replace(" ", "%20")
	city_dict["Weather"] = result['weather'][0]['description'].title()
	city_dict["Longitude"] = result['coord']['lon']
	city_dict["Latitude"] = result['coord']['lat']
	city_dict["Pressure"] = str(result['main']['pressure']) + " hPa"
	city_dict["Humidity"] = str(result['main']['humidity']) + "%"
	city_dict["Wind Speed"] = str(result['wind']['speed']) + "%"
	city_df = pd.DataFrame(city_dict.items(), columns = ["Metrics", "Values"])
	city_df = city_df.set_index("Metrics")

	return city_df

def excel_data(filename):
	""" Puts weather data in to excel file """
	wb = openpyxl.load_workbook(f'static/download_excel/{filename}') # Opens the excel file
	# Grabs sheet values and namees
	sheet1 = wb[wb.sheetnames[0]]
	sheet2 = wb[wb.sheetnames[1]]
	sheet3 = wb[wb.sheetnames[2]]
	sheet_name1 = wb.sheetnames[0]
	sheet_name2 = wb.sheetnames[1]
	sheet_name3 = wb.sheetnames[2]
	# Get city names
	city1 = sheet1['A1'].value
	city2 = sheet2['A1'].value
	# Gets data from API to a dataframe type
	city1_df = weather(city1)
	city2_df = weather(city2)
	# Merge the data from 2 cities 
	city3_df = pd.merge(city1_df, city2_df, left_index=True, right_on='Metrics')
	city3_df.columns = [city1, city2]
	# Writes to the excel file
	writer = pd.ExcelWriter(f'static/download_excel/{filename}')
	city1_df.to_excel(writer, sheet_name1)
	city2_df.to_excel(writer, sheet_name2)
	city3_df.to_excel(writer, sheet_name3)
	writer.save()
	
	wb = openpyxl.load_workbook(f'static/download_excel/{filename}')
	# Activates sheet 1
	wb.active = 0
	ws = wb.active
	# Writes chart to the excel sheet 1
	chart1 = BarChart()
	chart1.type = "col"
	chart1.style = 10
	chart1.title = f"Temperature in {city1}"
	chart1.y_axis.title = 'Temperature in Celsius'
	data1 = Reference(ws, min_col=2, min_row=2, max_row=5, max_col=2)
	cats1 = Reference(ws, min_col=1, min_row=2, max_row=5)
	chart1.add_data(data1, titles_from_data=False)
	chart1.set_categories(cats1)
	ws.add_chart(chart1, "D6")
	# Activates sheet 2
	wb.active = 1
	ws = wb.active
	# Writes chart to the excel sheet 2
	chart2 = BarChart()
	chart2.type = "col"
	chart2.style = 10
	chart2.title = f"Temperature in {city2}"
	chart2.y_axis.title = 'Temperature in Celsius'
	data2 = Reference(ws, min_col=2, min_row=2, max_row=5, max_col=2)
	cats2 = Reference(ws, min_col=1, min_row=2, max_row=5)
	chart2.add_data(data2, titles_from_data=False)
	chart2.set_categories(cats2)
	ws.add_chart(chart2, "D6")
	# Activates sheet 3
	wb.active = 2
	ws = wb.active
	# Writes chart to the excel sheet 3
	chart3 = BarChart()
	chart3.type = "col"
	chart3.style = 10
	chart3.title = f"Temperature in {city1} & {city2}"
	chart3.y_axis.title = 'Temperature in Celsius'
	data3 = Reference(ws, min_col=2, min_row=1, max_row=5, max_col=3)
	cats3 = Reference(ws, min_col=1, min_row=1, max_row=5)
	chart3.add_data(data3, titles_from_data=True)
	chart3.set_categories(cats3)
	ws.add_chart(chart3, "E6")
	# Saves the excel file
	wb.save(f'static/download_excel/{filename}')
	wb.close()

# Runs the website
if __name__ == '__main__':
	app.run(host='0.0.0.0', port=8080, debug = True)
